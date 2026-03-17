import openpyxl
import json
import datetime

wb = openpyxl.load_workbook(r'C:/Users/jarig/Desktop/Trainingszeiten angepasst.xlsx', data_only=True)
print("Sheets:", wb.sheetnames)
ws = wb.active

def fmt_time(val):
    if val is None:
        return ""
    if isinstance(val, datetime.time):
        return val.strftime("%H:%M")
    if isinstance(val, datetime.datetime):
        return val.strftime("%H:%M")
    # Excel numeric time fraction
    if isinstance(val, float):
        total_min = round(val * 24 * 60)
        h = total_min // 60
        m = total_min % 60
        return f"{h:02d}:{m:02d}"
    return str(val)

rows_out = []
for row in ws.iter_rows(min_row=3, values_only=False):
    r = [cell.value for cell in row]
    if not r[0]:
        continue
    sparte = str(r[0]).strip() if r[0] else ""
    gruppe = str(r[1]).strip() if r[1] else ""
    # Skip header rows where all cols are the same sparte name
    if sparte == gruppe and sparte != "":
        continue
    tag = str(r[2]).strip() if r[2] else ""
    uhr_von = fmt_time(r[3])
    uhr_bis = fmt_time(r[4])
    ort = str(r[5]).strip() if r[5] else ""
    jahreszeit = str(r[6]).strip() if r[6] else ""
    frequenz = str(r[7]).strip() if r[7] else ""
    trainer = str(r[8]).strip() if r[8] else ""
    email = str(r[9]).strip() if r[9] else ""
    telefon = str(r[10]).strip() if r[10] else ""

    if uhr_von and uhr_bis:
        uhrzeit = f"{uhr_von} - {uhr_bis}"
    elif uhr_von:
        uhrzeit = uhr_von
    else:
        uhrzeit = ""

    rows_out.append({
        "sparte": sparte,
        "gruppe": gruppe,
        "tag": tag,
        "uhrzeit": uhrzeit,
        "ort": ort,
        "jahreszeit": jahreszeit,
        "frequenz": frequenz,
        "trainer": trainer,
        "email": email,
        "telefon": telefon,
    })
    print(f"  {sparte} | {gruppe} | {tag} | {uhrzeit} | {ort}")

with open(r'C:/Users/jarig/Documents/sv-holm-seppensen/scripts/trainingszeiten-with-times.json', 'w', encoding='utf-8') as f:
    json.dump(rows_out, f, ensure_ascii=False, indent=2)

print(f"\nTotal: {len(rows_out)} Einträge gespeichert.")
